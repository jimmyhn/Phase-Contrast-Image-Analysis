# Imports: 
import os
import numpy as np
import torch
import torchvision
from PIL import Image
import matplotlib.pyplot as plt
from datetime import datetime
import re
import shutil
from skimage import measure
from skimage.segmentation import clear_border
import seaborn as sns
from scipy import stats
import pandas as pd
import sys
import io
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image as XLImage
import warnings
torchvision.disable_beta_transforms_warning() # Disable beta transforms warning when cellSAM is imported
from cellSAM import segment_cellular_image





# Functions:
def csAnalyze(img_array,img_name,quadrant=None):
    # Process images using CellSAM
    if bool(re.search(r"4x",img_name,re.IGNORECASE)):
            now = datetime.now()
            print(f"CellSAM started processing {quadrant} of image at {now.strftime('%H:%M:%S')}")
    else:
        now = datetime.now()
        print(f"CellSAM started processing image at {now.strftime('%H:%M:%S')}")

    try:
        
        img = np.clip(img_array // 256, 0, 255).astype(np.uint8) # Convert array from Uint16 to Uint8 for cellSAM input
        
        # Convert 2D to 3D
        img_3d = np.stack([img, img, img], axis=-1) # Stacks 3 of the 2D image (gray-scale) on top of each other making 3D image for cellSAM

        # Apply CellSAM
        device = torch.device('cuda') # Defines the device as being the GPU
        result = segment_cellular_image(img_3d, device=str(device)) # Segments the cells of the 3D image using the specified device
        # results = [mask, embedding, bounding_boxes]
        mask = result[0] # indexing only the mask to be saved aa a variable
        # The mask shows all cells identified by cellSAM, each cell is "shaded" by a unique integer value to be isolated

        # Convert to uint16
        mask = mask.astype(np.uint16) # Converted to ensure that each cell has it's own value (Uint8:256 vs Uint16: 65535 possible cell values)

        # Progress Tracker
        now = datetime.now()
        print(f"CellSAM finished processing at {now.strftime('%H:%M:%S')}")
        '''
        # Image Check
        plt.imshow(mask)
        plt.show() # forces plt to show image right away
        '''
        # Return mask back to main script
        return mask

    except Exception as e:
        print(f"It looks like {img_name} was not able to be processed because of this error:\n{str(e)}\nImage {img_name} will be skipped...")
        return None

def remove_outliers_IQR(data, mask=None):

    match type_of_outlier_filter:
        case 'bounds':
            interquartile_range = []
            deleted_data_index = None
            cell_values = None
            for x, kept in enumerate(data):
                if lowerBound <= kept['Area'] <= upperBound:
                    interquartile_range.append(kept['Area'])
                else:
                    if deleted_data_index is None:
                        deleted_data_index = []
                    if cell_values is None:
                        cell_values = []
                    deleted_data_index.append(x)
                    if mask is not None:
                        match kept['Quadrant']:
                            case 0:
                                mask[0]['P3mask'][np.isin(mask[0],kept['Area'])] = 0
                            case 1:
                                mask[1]['P3mask'][np.isin(mask[1],kept['Area'])] = 0
                            case 2:
                                mask[2]['P3mask'][np.isin(mask[2],kept['Area'])] = 0
                            case 3:
                                mask[3]['P3mask'][np.isin(mask[3],kept['Area'])] = 0
                    else:
                        cell_values.append(kept['Cell Value'])

        case 'percentile':
            Q_lower = np.percentile([d['Area'] for d in data], lower_percentile); print(f"\n{lower_percentile}th Quartile: {Q_lower}")
            Q_upper = np.percentile([d['Area'] for d in data], upper_percentile); print(f"{upper_percentile}th Quartile: {Q_upper}\n")
            interquartile_range = []
            deleted_data_index = None
            cell_values = None
            for x, kept in enumerate(data):
                if Q_lower <= kept['Area'] <= Q_upper:
                    interquartile_range.append(kept['Area'])
                else:
                    if deleted_data_index is None:
                        deleted_data_index = []
                    if cell_values is None:
                        cell_values = []
                    deleted_data_index.append(x)
                    if mask is not None:
                        match kept['Quadrant']:
                            case 0:
                                mask[0]['P3mask'][np.isin(mask[0],kept['Area'])] = 0
                            case 1:
                                mask[1]['P3mask'][np.isin(mask[1],kept['Area'])] = 0
                            case 2:
                                mask[2]['P3mask'][np.isin(mask[2],kept['Area'])] = 0
                            case 3:
                                mask[3]['P3mask'][np.isin(mask[3],kept['Area'])] = 0
                    else:
                        cell_values.append(kept['Cell Value'])


        case 'interquartile':
            Q1 = np.percentile([d['Area'] for d in data], 25); print(f"\n25th Quartile: {Q1}")
            Q3 = np.percentile([d['Area'] for d in data], 75); print(f"75th Quartile: {Q3}")
            IQR = Q3 - Q1
            lower_bound = Q1 - (1.5 * IQR); print(f"Lower Bound: {lower_bound}")
            upper_bound = Q3 + (1.5 * IQR); print(f"Upper Bound: {upper_bound}\n")
            interquartile_range = []
            deleted_data_index = None
            cell_values = None
            for x, kept in enumerate(data):
                if lower_bound <= kept['Area'] <= upper_bound:
                    interquartile_range.append(kept['Area'])
                else:
                    if deleted_data_index is None:
                        deleted_data_index = []
                    if cell_values is None:
                        cell_values = []
                    deleted_data_index.append(x)
                    if mask is not None:
                        match kept['Quadrant']:
                            case 0:
                                mask[0]['P3mask'][np.isin(mask[0],kept['Area'])] = 0
                            case 1:
                                mask[1]['P3mask'][np.isin(mask[1],kept['Area'])] = 0
                            case 2:
                                mask[2]['P3mask'][np.isin(mask[2],kept['Area'])] = 0
                            case 3:
                                mask[3]['P3mask'][np.isin(mask[3],kept['Area'])] = 0
                    else:
                        cell_values.append(kept['Cell Value'])


    # print(interquartile_range)
    return interquartile_range, cell_values, deleted_data_index, mask

def normal_distribution_test(data):
    if len(data) > 3:
        # Sharpio-Wilk Test
        statistic, p_value = stats.shapiro(data)
    else:
        statistic = np.nan
        p_value = np.nan
        

    # Descriptive Statistics
    mean = round(np.mean(data), 1)
    median = np.median(data)
    mode = stats.mode(data).mode

    # Skewness
    skewness = round(stats.skew(data),3)

    test_results = {
        'Sharpio Statistic' : statistic,
        'Sharpio P-Value' : p_value,
        'Skewness Statistic' : skewness,
        'Mean' : mean,
        'Median' : median,
        'Mode' : mode
    }
    
    return test_results

def combine_quadrants(quad,boolean=False,map_type='gray'):
    top_half = np.hstack([quad[0],quad[1]])
    bottom_half= np.hstack([quad[2],quad[3]])
    mask = np.vstack([top_half,bottom_half])
    if boolean:
        plt.imshow(mask,cmap = map_type)
        plt.show()

    return mask

def saveArray(array, filename, folderPath, sampleFolder, formatType):
    createdPath = os.path.join(sampleFolder, folderPath)
    os.makedirs(createdPath,exist_ok=True)
    ext = '.' + formatType.lower()
    if ext == '.tiff':
        ext = '.tif'
    elif ext == '.jpg':
        ext = '.jpeg'

    path_of_file =  os.path.join(createdPath,filename + ext)
    img = Image.fromarray(array)
    img.save(path_of_file,format=formatType)

class ImageFileWarning(Warning):
    pass



# MAIN Script:

# Establshing Constants and Symbols:
micro = "\u03BC"
squared = "\u00B2"

conversion4x = 1.6125
conversion10x = 0.645
conversion20x = 0.3225

GY_threshold = 37500 # pixel value of when they are green/yellow
filter_threshold = 0.40 # the percentage of how much green/yellow there needs to before it considered not a cell

user = "Nasim"
images_in_sheets = True
type_of_outlier_filter = 'percentile' # choose 'bounds', 'interquartile', 'percentile'
lowerBound = 200
upperBound = 1000

lower_percentile= 10
upper_percentile= 90

starting_index = 0
inputFolder = 'Input'

# Navigate to the input folder
main_folder = os.getcwd(); print(f"Main dir: {main_folder}")
analysis_folder = os.path.join(main_folder, inputFolder); print(f"Input dir: {analysis_folder}")
print(f"Currently in main folder: {os.listdir()}")
os.chdir(analysis_folder); print(f"Changes dir to input folder: {os.listdir()}") 

# List the "types" of images that will be analyzed
types_to_analyze = [f for f in os.listdir() if os.path.isdir(f)] # creates a list of folders within current directory
if '.ipynb_checkpoints' in types_to_analyze:
        types_to_analyze.remove('.ipynb_checkpoints')

# Loops through the horizontal and vertical folders
for a in types_to_analyze:
    match a:
        case "Horizontal":
            is_Vertical_Orientation = False

        case "Vertical":
            is_Vertical_Orientation = True

    # Sorting/Listing out all the images that will be analyzed
    print(f"Now going through {a} type images...")
    os.chdir(f"./{a}"); print (f"Current dir is: {os.path.basename(os.getcwd())}")

    # Get list of image sets to analyze
    folders_to_analyze = [f for f in os.listdir() if os.path.isdir(f)]
    if '.ipynb_checkpoints' in folders_to_analyze:
        folders_to_analyze.remove('.ipynb_checkpoints')
        
    print(f"Folders to analyze:{folders_to_analyze}")

    # Get list of image sets' directories that will be analyzed
    folders_to_analyze_dir = [os.path.join(os.getcwd(),f) for f in os.listdir() if os.path.isdir(f)]

    # Collect all images from subfolders
    for f, folder in enumerate(folders_to_analyze): # first for loop goes through each sample
        # Initialize list to store image information
        analysisReference = [] # this will be where everything is saved to used throughout the code

        hours_folder = os.path.join(analysis_folder, a, folder)
        print(f"Now analyzing {folder} sample")
        for root, _, files in os.walk(hours_folder): # second for loop goes through each hours folder
            # Accounting for folders that are not hour folders
            if bool(re.search(r"hrs|hr|hours|hour",os.path.basename(root),re.IGNORECASE)):
                for file in files: # third for loop goes through each image
                    if file.lower().endswith(('.tif')): # filters in only the .tif files
                        image_path = os.path.join(root, file)
                        # An exception just in case file can't be opened
                        try :
                            img = np.array(Image.open(image_path)) # opens the image as an array
                        except Exception as e:
                            previous_dir = os.getcwd()
                            os.chdir(os.path.basename(hours_folder))
                            os.makedirs('Images_Not_Analyzed', exist_ok=True) # checks to see folder already exist, if not then creates it
                            try:
                                shutil.copy2(os.path.join(root,file),os.path.join(os.getcwd(),'Images_Not_Analyzed', file))
                            except:
                                os.chdir(previous_dir)
                                print(f"It looks like {file} was not able to be open because of this error:\n{str(e)}\nImage {file} will be skipped nor saved on 'Images Not Analyzed'...")
                                continue
                            os.chdir(previous_dir)
                            print(f"It looks like {file} was not able to be open because of this error:\n{str(e)}\nImage {file} will be skipped but saved on 'Images Not Analyzed'...")
                            continue

                        if len(img.shape) == 2:  # filters in only  grayscale images
                            if user == "Nasim":
                                analysisReference.append({
                                    'originalName': file,
                                    'name': os.path.basename(root) + "_" + file,
                                    'folder': root
                                })
                            else: 
                                analysisReference.append({
                                    'name': file,
                                    'folder': root
                                })
        print(f"TIFF Images in current folder:\n{[item['name'] for item in analysisReference]}\n")


        # Image Processing Starts Here
        if f == 0:
            number_of_slides_done = 0 + starting_index
            number_of_slides_to_analyze = len(analysisReference)
            analysisReference = analysisReference[starting_index:]
            if user == '': #empty if
                b = []
                for image in analysisReference:
                    if bool(re.search(r"10x",image['name'],re.IGNORECASE)):
                        if bool(re.search(r"Center",image['name'],re.IGNORECASE)):
                            b.append(image)
                analysisReference = b
                number_of_slides_to_analyze = len(analysisReference)
                number_of_slides_done = 0

        else:
            number_of_slides_done = 0
            number_of_slides_to_analyze = len(analysisReference)
            if user == '':  #empty if
                b = []
                for image in analysisReference:
                    if bool(re.search(r"10x",image['name'],re.IGNORECASE)):
                        if bool(re.search(r"Center",image['name'],re.IGNORECASE)):
                            b.append(image)
                analysisReference = b
                number_of_slides_to_analyze = len(analysisReference)
                number_of_slides_done = 0

        for image_index, image in enumerate(analysisReference): # goes through each image
            if f == 0:
                image_index = image_index + starting_index 

            now = datetime.now()
            print(f"Image processing for {image['name']} started at {now.strftime('%H:%M:%S')}\n")
            
            if user == "Nasim":
                array1 = np.array (Image.open(os.path.join(image['folder'],image['originalName']))) # opens the image as an array
            else:
                array1 = np.array (Image.open(os.path.join(image['folder'],image['name']))) # opens the image as an array


            if array1.dtype != np.uint16: # a check just in case images are formatted wrong on Dreamscope
                raise TypeError(f"The image is {array1.dtype} type, it should be Uint16. Please filter it out before running it again")
            
            # Rotation Process (if vertical type):
            if is_Vertical_Orientation:
                array2 = np.rot90(array1)

                # Normilaztion Process in the Vertical Case:
                # Normalizing phase brightness, so that all images will have the same brightness intensity 
                # Honestly, I should take the oppurtunity to see whether this brightness step is making the analysis more or less accurate
                average_brightness = array2.mean(); # finds the average brightness value of image (float value)
                wanted_brightness = 35000 # the brightness values that you want
                array2 = (array2.astype(np.float64)/average_brightness) * wanted_brightness
                array2 = np.clip(array2, 0, 65535)
                array2 = array2.astype(np.uint16)

            else: 
                # Normilaztion Process in the Horizontal Case:
                # Normalizing phase brightness, so that all images will have the same brightness intensity 
                # Honestly, I should take the oppurtunity to see whether this brightness step is making the analysis more or less accurate
                average_brightness = array1.mean() # finds the average brightness value of image (float value)
                wanted_brightness = 35000 # the brightness values that you want
                array2 =  (array1.astype(np.float64)/average_brightness) * wanted_brightness
                array2 = array2.astype(np.uint16)


            # Magnification Sorting
            if bool(re.search(r"4x",image['name'],re.IGNORECASE)):

                # Quadrant Splitting Process:
                # spliting 4x images into 4 quadrants so that cellSAM can identify cells more accurately
                print(f"Now analying {image['name']}")
                quadrants = []
                width_split = round(array2.shape[1]/2)
                height_split = round(array2.shape[0]/2)
                array_TL = array2[:height_split,:width_split] # top left
                array_TR = array2[:height_split,width_split:] # top right
                array_BL = array2[height_split:,:width_split] # bottom left
                array_BR = array2[height_split:,width_split:] # bottom right
                quadrants.extend([
                    {'name': "Quadrant 1",'array' : array_TL},
                    {'name': "Quadrant 2",'array' : array_TR},
                    {'name': "Quadrant 3",'array' : array_BL},
                    {'name': "Quadrant 4",'array' : array_BR}
                ])


                # Image Analysis Starts Here:
                total_OG_cells = 0
                total_filtered_cells_P1 = 0
                total_filtered_cells_P2 = 0
                subtotals = []
                raw_data = []
                raw_data_for_P3mask = []
                raw_data_outlier_P2 = []
                all_quadrants_processed = True
                for quadrant_index, quadrant in enumerate(quadrants): # loops through each quadrant: Q1,Q2,Q3,Q4 image


                    # Creation of Masks
                    array = quadrant['array']
                    previous_dir = os.getcwd()
                    os.chdir(os.path.basename(hours_folder))
                    if os.path.isdir(os.path.join(os.getcwd(),"Unprocessed_Masks")):
                        for root, _, files in os.walk(os.path.join(os.getcwd(),"Unprocessed_Masks")):
                            unprocessed_masks = [{
                                'Filepath': root,
                                'Filename': files
                            }]
                        mask_files = unprocessed_masks[0]['Filename']
                        if len(mask_files) > 0:
                            mask_file = next((f for f in mask_files if f.startswith(os.path.splitext(image['name'])[0]+"_"+ quadrant['name'])),None)
                            if mask_file is not None:
                                mask = np.array(Image.open(os.path.join(unprocessed_masks[0]['Filepath'],mask_file)))
                            else:
                                mask = csAnalyze(array,image['name'],quadrant['name'])
                        else:
                            mask = csAnalyze(array,image['name'],quadrant['name'])
                    else:
                        mask = csAnalyze(array,image['name'],quadrant['name']) # cellSAM function is used right here

                    os.chdir(previous_dir)
                        
                    if mask is None: # in case cellSAM is not able to process the image, this will the image analysis for this whole image
                        all_quadrants_processed = False
                        break
                
                    # Differentiating cells within mask 
                    unique_mask_values = np.unique(mask)
                    unique_mask_values = unique_mask_values[unique_mask_values != 0]

                    print(f"The amount of cells cellSAM orignally found in {quadrant['name']}: {unique_mask_values.size}")
                    total_OG_cells += unique_mask_values.size

                    # Saving Original 4x Masks as an extra precaution
                    saveArray(mask, os.path.splitext(image['name'])[0] + '_' + quadrant['name'] + '_unprocessedMask', 'Unprocessed_Masks', hours_folder, 'TIFF')

                    quadrants[quadrant_index]['mask'] = mask
                    quadrants[quadrant_index]['boolean_mask'] = (mask>0)*1 # stores the logical array mask for saving purposes
                    # Phase 1 Filtering: Edge Exclusion Process
                    # Cells that are touching the border will be "deleted"
                    filtered_mask = clear_border(mask)
                    unique_mask_values = np.unique(filtered_mask)
                    unique_mask_values = unique_mask_values[unique_mask_values != 0]; print(f"P1 Subtotal for {quadrant['name']}: {unique_mask_values.size}")
                    total_filtered_cells_P1 += unique_mask_values.size
                    quadrants[quadrant_index]['filtered_boolean_mask'] = (filtered_mask>0)*1 # stores the logical filtered array mask for saving purposes



                    # Phase 2 Filtering: Pixel Value Filter (for white spots)
                    fake_cells = 0
                    cells_to_remove_P2 = [] # resets removal list P2 every quadrant loop
                    for unique_value in (unique_mask_values): # loops through each cell
                        cell_mask = (filtered_mask == unique_value)*1 # the *1 is to covert logic array to number array
                        cell_image_array = array*cell_mask
                        unique_cell_pixels,counts =  np.unique(cell_image_array,return_counts=True)
                        # Find the index of the pixel values closest to the threshold
                        start_index = np.searchsorted(unique_cell_pixels, GY_threshold)
                        end_index = np.searchsorted(unique_cell_pixels,65535,side='right') # 65535 is the max value which yellow
                        
                        unwanted_pixel = sum(counts[start_index:end_index])
                        total_pixel = sum(counts[1:end_index]) # 1 because the majority of the image is 0
                        unwanted_to_total = unwanted_pixel/total_pixel


                        # if the cell is "acceptable" it will pass onto analysis, if not, then it will be skipped/filtered out
                        if unwanted_to_total < filter_threshold: 

                            # Analysis of Each Cell Starts Here:
                            regions = measure.regionprops(cell_mask) # can regionprops in one go, the whole mask; might do it if filter isn't needed/scraped
                            region = regions[0] # have to do this because its a list
                            if region.perimeter == 0 or region.axis_minor_length == 0:
                                circularity = np.nan
                                polarity = np.nan
                            else: 
                                circularity = ((2*np.sqrt(np.pi*region.area))/region.perimeter)
                                polarity = (region.axis_major_length/region.axis_minor_length)

                            # Raw data collection
                            raw_data.append({
                                'Orientation(deg):': np.degrees(region.orientation),
                                'Circularity:': circularity,
                                'Polarity:': polarity,
                                'Eccentricity:': region.eccentricity,
                                f'Major Length({micro}m):': region.axis_major_length * conversion4x, 
                                f'Minor Length({micro}m):': region.axis_minor_length * conversion4x, 
                                f'Perimeter({micro}m):': region.perimeter * conversion4x, 
                                f'Area({micro}m{squared}):': region.area * (conversion4x ** 2)
                            })
                            
                            raw_data_for_P3mask.append({
                                'Area': region.area * (conversion4x ** 2),
                                'Quadrant' : quadrant_index,
                                'Cell Value': unique_value
                            })

                        else:
                            # print(f"Cell {unique_value} holds {unwanted_to_total*100}% of green/yellow pixels, so it's fake")
                            fake_cells += 1
                            cells_to_remove_P2.append(unique_value) # used for other outputs including outliers and visuals

                    # Filtered Cell Total Calculations:
                    subtotal_cells = unique_mask_values.size - fake_cells # calculates the total amount of filtered cells left in one quadrant
                    subtotals.append(subtotal_cells)
                    print(f"Out of the {unique_mask_values.size} cells, pixel value filtering has determined {fake_cells} were not cells\nThere is now a total of {subtotal_cells} cells in this qudrant")
                    
                    total_filtered_cells_P2 += subtotal_cells # calculates the total amount of filtered cells left in the whole image
                    print(f"P2 Subtotal for {quadrant['name']}: {subtotal_cells}")


                    # Measuring Outlier Raw Data of Phase 2
                    if cells_to_remove_P2 != []:
                        for unique_value in cells_to_remove_P2:
                            cell_mask = (filtered_mask == unique_value)*1
                            regions = measure.regionprops(cell_mask)
                            region = regions[0] # have to do this because its a list
                            if region.perimeter == 0 or region.axis_minor_length == 0:
                                circularity = np.nan
                                polarity = np.nan
                            else: 
                                circularity = ((2*np.sqrt(np.pi*region.area))/region.perimeter)
                                polarity = (region.axis_major_length/region.axis_minor_length)

                            # Raw data collection
                            raw_data_outlier_P2.append({
                                'Orientation(deg):': np.degrees(region.orientation),
                                'Circularity:': circularity,
                                'Polarity:': polarity,
                                'Eccentricity:': region.eccentricity,
                                f'Major Length({micro}m):': region.axis_major_length * conversion4x, 
                                f'Minor Length({micro}m):': region.axis_minor_length * conversion4x, 
                                f'Perimeter({micro}m):': region.perimeter * conversion4x, 
                                f'Area({micro}m{squared}):': region.area * (conversion4x ** 2)
                            })

                        P2mask = filtered_mask
                        P2mask[np.isin(P2mask,cells_to_remove_P2)] = 0

                    else:
                        # Creating Image after phase 2 filtering
                        P2mask = filtered_mask


                    quadrants[quadrant_index]['P3mask'] = P2mask

                    P2mask = (P2mask>0)*1
                    P2quadrant = quadrant['array'] * P2mask

                    quadrants[quadrant_index]['P2image'] = P2quadrant


                    
                # A check to see if quadrants are being processed correctly at this point
                if not all_quadrants_processed: # in case of cellSAM not being able to process the image
                    previous_dir = os.getcwd()
                    os.chdir(os.path.basename(hours_folder))
                    print(os.getcwd())
                    os.makedirs('Images_Not_Analyzed', exist_ok=True) # checks to see folder already exist, if not then creates it
                    shutil.copy2(os.path.join(image['folder'],image['originalName']),os.path.join(os.getcwd(),'Images_Not_Analyzed', image['name']))
                    os.chdir(previous_dir)
                    warnings.warn(f"{image['name']} was not able to be analyzed. Please check why in 'Images_Not_Analyzed' folder...", ImageFileWarning)
                    continue
                print(f"P1 total for {image['name']}: {total_filtered_cells_P1}")
                print(f"P2 total for {image['name']}: {total_filtered_cells_P2}") 

                # Creating Series for Raw Data Outliers of P2:
                raw_data_outlier_P2 = pd.DataFrame(raw_data_outlier_P2)

                # Phase 3 Filtering: Outlier Filter for P3 Raw Data
                normal_distribution_data, _ , cells_to_remove_P3, quadrants = remove_outliers_IQR(raw_data_for_P3mask,quadrants)

                if cells_to_remove_P3 is not None:
                    print(f"Cells removed in P3 filter: {len(cells_to_remove_P3)}")
                else:
                    print(f"Cells removed in P3 filter: 0")

                for quadrant_index in range(4):
                    quadrants[quadrant_index]['boolean_P3mask'] = (quadrants[quadrant_index]['P3mask']>0)*1

                # Creating the last P3 Image after filtering
                P3mask = combine_quadrants([q['boolean_P3mask'] for q in quadrants])
                whole_image_P3 = array2*P3mask

                # Area Data for Initial Graphs
                area = [element[f'Area({micro}m{squared}):'] for element in raw_data]

                # Visual Results For 4x:
                # CellSAM Mask of the whole 4x image
                whole_mask = combine_quadrants([q['mask'] for q in quadrants])  # potential delete

                # Boolean Mask of the whole 4x image
                whole_boolean_mask = combine_quadrants([q['boolean_mask'] for q in quadrants])
                
                # Boolean Filtered Mask P1 of the whole 4x image
                whole_boolean_mask_filteredP1 = combine_quadrants([q['filtered_boolean_mask'] for q in quadrants])

                # Whole image of 4x (Phase 2)
                whole_image_P2 = combine_quadrants([q['P2image'] for q in quadrants])

                        
            elif bool(re.search(r"10x",image['name'],re.IGNORECASE)) | bool(re.search(r"20x",image['name'],re.IGNORECASE)):
                
                if bool(re.search(r"10x",image['name'],re.IGNORECASE)):
                    b = '10x'
                elif bool(re.search(r"20x",image['name'],re.IGNORECASE)):
                    b ='20x'
                else:
                    previous_dir = os.getcwd()
                    os.chdir(os.path.basename(hours_folder))
                    print(os.getcwd())
                    os.makedirs('Images_Not_Analyzed', exist_ok=True) # checks to see folder already exist, if not then creates it
                    shutil.copy2(os.path.join(image['folder'],image['originalName']),os.path.join(os.getcwd(),'Images_Not_Analyzed', image['name']))
                    os.chdir(previous_dir)
                    warnings.warn(f"{image['name']} was not able to be analyzed. Please check why in 'Images_Not_Analyzed' folder...", ImageFileWarning)
                    continue

                print(f"Now analying {image['name']}")

                # Image Analysis for 10x/20x Starts Here:
                total_OG_cells = 0
                total_filtered_cells_P1 = 0
                total_filtered_cells_P2 = 0
                raw_data = []
                raw_data_for_P3mask = []
                raw_data_outlier_P2 = []

                # Creation of Masks
                previous_dir = os.getcwd()
                os.chdir(os.path.basename(hours_folder))
                if os.path.isdir(os.path.join(os.getcwd(),"Unprocessed_Masks")):
                    for root, _, files in os.walk(os.path.join(os.getcwd(),"Unprocessed_Masks")):
                        unprocessed_masks = [{
                            'Filepath': root,
                            'Filename': files
                        }]
                    mask_files = unprocessed_masks[0]['Filename']
                    if len(mask_files) > 0:
                        mask_file = next((f for f in mask_files if f.startswith(os.path.splitext(image['name'])[0])),None)
                        if mask_file is not None:
                            mask = np.array(Image.open(os.path.join(unprocessed_masks[0]['Filepath'],mask_file)))
                        else:
                            mask = csAnalyze(array2,image['name'])
                    else:
                        mask = csAnalyze(array2,image['name'])
                else:
                    mask = csAnalyze(array2,image['name']) # cellSAM function is used right here

                os.chdir(previous_dir)

                    

 
                if mask is None: # in case cellSAM is not able to process the image, this will skip the image analysis for this whole image
                    previous_dir = os.getcwd()
                    os.chdir(os.path.basename(hours_folder))
                    print(os.getcwd())
                    os.makedirs('Images_Not_Analyzed', exist_ok=True) # checks to see folder already exist, if not then creates it
                    shutil.copy2(os.path.join(image['folder'],image['originalName']),os.path.join(os.getcwd(),'Images_Not_Analyzed', image['name']))
                    os.chdir(previous_dir)
                    warnings.warn(f"{image['name']} was not able to be analyzed. Please check why in 'Images_Not_Analyzed' folder...", ImageFileWarning)
                    continue
                
                # Differentiating cells within mask 
                unique_mask_values = np.unique(mask)
                unique_mask_values = unique_mask_values[unique_mask_values != 0]

                print(f"The amount of cells cellSAM orignally found in {image['name']}: {unique_mask_values.size}")
                total_OG_cells = unique_mask_values.size

                # Saving Original Masks as an extra precaution
                saveArray(mask, os.path.splitext(image['name'])[0] + '_unprocessedMask', 'Unprocessed_Masks', hours_folder, 'TIFF')

                whole_mask = mask
                whole_boolean_mask = (mask>0)*1 # stores the logical array mask for saving purposes
                


                # Phase 1 Filtering: Edge Exclusion Process
                # Cells that are touching the border will be "deleted"
                filtered_mask = clear_border(mask)
                unique_mask_values = np.unique(filtered_mask)
                unique_mask_values = unique_mask_values[unique_mask_values != 0]; print(f"P1 total for {image['name']}: {unique_mask_values.size}")
                total_filtered_cells_P1 = unique_mask_values.size
                whole_boolean_mask_filteredP1 = (filtered_mask>0)*1 # stores the logical filtered array mask for saving purposes

                # Phase 2 Filtering: Pixel Value Filter (for white spots)
                fake_cells = 0
                cells_to_remove_P2 = [] # resets removal list P2 every quadrant loop
                for unique_value in (unique_mask_values): # loops through each cell
                    cell_mask = (filtered_mask == unique_value)*1 # the *1 is to covert logic array to number array
                    cell_image_array = array2*cell_mask
                    unique_cell_pixels,counts =  np.unique(cell_image_array,return_counts=True)
                    # Find the index of the pixel values closest to the threshold
                    start_index = np.searchsorted(unique_cell_pixels, GY_threshold)
                    end_index = np.searchsorted(unique_cell_pixels,65535,side='right') # 65535 is the max value which yellow
                    
                    unwanted_pixel = sum(counts[start_index:end_index])
                    total_pixel = sum(counts[1:end_index]) # 1 because the majority of the image is 0
                    unwanted_to_total = unwanted_pixel/total_pixel


                    # if the cell is "acceptable" it will pass onto analysis, if not, then it will be skipped/filtered out
                    if unwanted_to_total < filter_threshold: 

                        # Analysis of Each Cell Starts Here:
                        regions = measure.regionprops(cell_mask) # can regionprops in one go, the whole mask; might do it if filter isn't needed/scraped
                        region = regions[0] # have to do this because its a list
                        if region.perimeter == 0 or region.axis_minor_length == 0:
                            circularity = np.nan
                            polarity = np.nan
                        else: 
                            circularity = ((2*np.sqrt(np.pi*region.area))/region.perimeter)
                            polarity = (region.axis_major_length/region.axis_minor_length)

                        # Raw data collection
                        match b:
                            case '10x': 
                                raw_data.append({
                                    'Orientation(deg):': np.degrees(region.orientation),
                                    'Circularity:': circularity,
                                    'Polarity:': polarity,
                                    'Eccentricity:': region.eccentricity,
                                    f'Major Length({micro}m):': region.axis_major_length * conversion10x, 
                                    f'Minor Length({micro}m):': region.axis_minor_length * conversion10x, 
                                    f'Perimeter({micro}m):': region.perimeter * conversion10x, 
                                    f'Area({micro}m{squared}):': region.area * (conversion10x ** 2)
                                })
                                raw_data_for_P3mask.append({
                                'Area': region.area * (conversion10x ** 2),
                                'Cell Value': unique_value
                                })      
                            case '20x':
                                raw_data.append({
                                    'Orientation(deg):': np.degrees(region.orientation),
                                    'Circularity:': circularity,
                                    'Polarity:': polarity,
                                    'Eccentricity:': region.eccentricity,
                                    f'Major Length({micro}m):': region.axis_major_length * conversion20x, 
                                    f'Minor Length({micro}m):': region.axis_minor_length * conversion20x, 
                                    f'Perimeter({micro}m):': region.perimeter * conversion20x, 
                                    f'Area({micro}m{squared}):': region.area * (conversion20x ** 2)
                                })
                                raw_data_for_P3mask.append({
                                'Area': region.area * (conversion20x ** 2),
                                'Cell Value': unique_value
                                }) 
                    else:
                        # print(f"Cell {unique_value} holds {unwanted_to_total*100}% of green/yellow pixels, so it's fake")
                        fake_cells += 1
                        cells_to_remove_P2.append(unique_value) # used for other outputs including outliers and visuals

                # Filtered Cell Total Calculations:
                print(f"Out of the {unique_mask_values.size} cells, pixel value filtering has determined {fake_cells} were not cells...")
                total_filtered_cells_P2 = unique_mask_values.size - fake_cells # calculates the total amount of filtered cells left in the whole image
                print(f"P2 Total for {image['name']}: {total_filtered_cells_P2}")


                # Measuring Outlier Raw Data of Phase 2
                if cells_to_remove_P2 != []:
                    for unique_value in cells_to_remove_P2:
                        cell_mask = (filtered_mask == unique_value)*1
                        regions = measure.regionprops(cell_mask)
                        region = regions[0] # have to do this because its a list
                        if region.perimeter == 0 or region.axis_minor_length == 0:
                            circularity = np.nan
                            polarity = np.nan
                        else: 
                            circularity = ((2*np.sqrt(np.pi*region.area))/region.perimeter)
                            polarity = (region.axis_major_length/region.axis_minor_length)

                        # Raw outlier data collection (Phase2)
                        match b:
                            case '10x': 
                                raw_data_outlier_P2.append({
                                    'Orientation(deg):': np.degrees(region.orientation),
                                    'Circularity:': circularity,
                                    'Polarity:': polarity,
                                    'Eccentricity:': region.eccentricity,
                                    f'Major Length({micro}m):': region.axis_major_length * conversion10x, 
                                    f'Minor Length({micro}m):': region.axis_minor_length * conversion10x, 
                                    f'Perimeter({micro}m):': region.perimeter * conversion10x, 
                                    f'Area({micro}m{squared}):': region.area * (conversion10x ** 2)
                                })
                            case '20x':
                                raw_data_outlier_P2.append({
                                    'Orientation(deg):': np.degrees(region.orientation),
                                    'Circularity:': circularity,
                                    'Polarity:': polarity,
                                    'Eccentricity:': region.eccentricity,
                                    f'Major Length({micro}m):': region.axis_major_length * conversion20x, 
                                    f'Minor Length({micro}m):': region.axis_minor_length * conversion20x, 
                                    f'Perimeter({micro}m):': region.perimeter * conversion20x, 
                                    f'Area({micro}m{squared}):': region.area * (conversion20x ** 2)
                                })

                    # Creating Image after phase 2 filtering
                    P2mask = filtered_mask
                    P2mask[np.isin(P2mask,cells_to_remove_P2)] = 0
                else:
                    P2mask = filtered_mask

                P2mask_boolean = (P2mask>0)*1
                whole_image_P2 = array2 * P2mask_boolean

                # Creating Series for Raw Data Outliers of P2:
                raw_data_outlier_P2 = pd.DataFrame(raw_data_outlier_P2)

                # Phase 3 Filtering: Outlier Filter for P3 (for small black spots)
                # Interquartile Method For 10x/20x
                normal_distribution_data, cell_value_to_remove_P3, cells_to_remove_P3, _  = remove_outliers_IQR(raw_data_for_P3mask)

                if cells_to_remove_P3 is not None:
                    print(f"Cells removed in P3 filter: {len(cells_to_remove_P3)}")
                else:
                    print(f"Cells removed in P3 filter: 0")


                # Creating the last P3 Image after filtering
                P3mask = P2mask
                if cell_value_to_remove_P3 is not None:
                    for unique_value in cell_value_to_remove_P3:
                        P3mask[np.isin(P3mask,unique_value)] = 0
                P3mask = (P3mask>0)*1
                whole_image_P3 = array2*P3mask
                
                # Area Data for Initial Graphs
                area = [element[f'Area({micro}m{squared}):'] for element in raw_data]

            else: # A check in case any files do not have 4x, 10x, 20x on file
                previous_dir = os.getcwd()
                os.chdir(os.path.basename(hours_folder))
                os.makedirs('Images_Not_Analyzed', exist_ok=True) # checks to see folder already exist, if not then creates it
                shutil.copy2(os.path.join(image['folder'],image['originalName']),os.path.join(os.getcwd(),'Images_Not_Analyzed', image['name']))
                os.chdir(previous_dir)
                warnings.warn(f"{image['name']} was not able to be analyzed. Please check why in 'Images_Not_Analyzed' folder...", ImageFileWarning)
                continue
            
            # Visual Results Continued for all Magnifications:
            # Normal Distribution Tests (Before):
            test_before_results = normal_distribution_test(area)

            # Normal Distribution Tests (After):
            test_after_results = normal_distribution_test(normal_distribution_data)

            # Histograms and Box & Whiskers plots
            if images_in_sheets:
                fig, ((ax1, ax2), (ax3, ax4)) = plt.subplots(2, 2, figsize=(16,9))

                # Subplot 1: Original Histogram
                ax1.hist(area, bins=30, edgecolor='black')
                ax1.set_title('Histogram of Area (Original)')
                ax1.set_xlabel(f'Area({micro}m{squared})')
                ax1.set_ylabel('Frequency')

                # Subplot 2: Original Box & Whiskers
                sns.boxplot(x=['']*len(area), y=area, hue=['']*len(area), ax=ax2, palette="Set2", legend=False)
                sns.swarmplot(data=area, color="red", size=1, alpha=0.7, ax=ax2)
                ax2.set_title('Original Data')
                ax2.set_ylabel(f'Area({micro}m{squared})')

                # Subplot 3: Normal Histogram
                ax3.hist(normal_distribution_data, bins=30, edgecolor='black')
                ax3.set_title('Histogram of Area (Normal)')
                ax3.set_xlabel(f'Area({micro}m{squared})')
                ax3.set_ylabel('Frequency')

                # Subplot 4: Normal Box & Whiskers
                sns.boxplot(x=['']*len(normal_distribution_data), y=normal_distribution_data, hue=['']*len(normal_distribution_data), ax=ax4, palette="Set2", legend=False)
                sns.swarmplot(data=normal_distribution_data, color="red", size=1, alpha=0.7, ax=ax4)
                ax4.set_title('Normal Distributed Data')
                ax4.set_ylabel(f'Area({micro}m{squared})')

                # Adjust layout and save into variable
                plt.tight_layout()
                normal_distribution_bufferFig = io.BytesIO()
                fig.savefig(normal_distribution_bufferFig, format='png', dpi= 600, bbox_inches='tight')
                normal_distribution_bufferFig.seek(0)
                plt.close(fig)


            # Storing outlier raw data of phase 3 filtering:
            if cells_to_remove_P3 is not None:
                raw_data = pd.DataFrame(raw_data)
                raw_data_outlier_P3 = raw_data.iloc[cells_to_remove_P3].reset_index(drop=True)
                # Getting rid of outliers in the main data set
                cells_to_remove_P3_labels = raw_data.index[cells_to_remove_P3]
                raw_data = raw_data.drop(cells_to_remove_P3_labels).reset_index(drop=True)
            else:
                raw_data = pd.DataFrame(raw_data)
                raw_data_outlier_P3 = pd.DataFrame()

            total_filtered_cells_P3 = len(raw_data)

            print(f"After phase 3 filtering, the final cell count is: {total_filtered_cells_P3}")
            

            # Mask with Adjusted Image        
            masked_image = array2*whole_boolean_mask

            # P1Mask with Adjusted Image
            masked_filteredP1_image = array2 * whole_boolean_mask_filteredP1
            
            # Image Process Figure Creation
            if images_in_sheets:
                fig, axs = plt.subplots(2,4,figsize=(16,9))
                axs = axs.flatten()

                # 1. Original Image
                axs[0].imshow(array1,cmap='gray')
                axs[0].set_title('Original Image')
                axs[0].axis('off')

                # 2. Normalization Image
                axs[1].imshow(array2,cmap='gray')
                axs[1].set_title('Normalization Image')
                axs[1].axis('off')

                # 3. CellSAM Mask
                axs[2].imshow(whole_mask,cmap='viridis')
                axs[2].set_title('CellSAM Mask')
                axs[2].axis('off')

                # 4. Combined Masked Image
                axs[3].imshow(masked_image,cmap='gray')
                axs[3].set_title('Combined Masked Image')
                axs[3].axis('off')

                # 5. Edge Filtered Image (Phase 1)
                axs[4].imshow(masked_filteredP1_image,cmap='gray')
                axs[4].set_title('Edge Filtered Image(P1)')
                axs[4].axis('off')

                # 6. Viridis Color Map Image (Phase 1)
                axs[5].imshow(masked_filteredP1_image,cmap='viridis')
                axs[5].set_title('Viridis Color Map Image(P1)')
                axs[5].axis('off')

                # 7. Pixel Value Filtered Image (Phase 2)
                axs[6].imshow(whole_image_P2,cmap='viridis')
                axs[6].set_title('Pixel Value Filtered Image(P2)')
                axs[6].axis('off')

                # 8. Normal Distribution Filtered Image (Phase 3)
                axs[7].imshow(whole_image_P3,cmap='gray')
                axs[7].set_title('Normal Distribution Filtered Image(P3)')
                axs[7].axis('off')

                plt.tight_layout()
                img_processing_bufferFig = io.BytesIO()
                fig.savefig(img_processing_bufferFig, format='png', dpi= 600, bbox_inches='tight')
                img_processing_bufferFig.seek(0)
                plt.close(fig)

            now = datetime.now()
            print(f"Image processing for {image['name']} has been completed at {now.strftime('%H:%M:%S')}")



            # Spreadsheets
            createdPath = os.path.join(hours_folder, f"Raw_Data_Spreadsheets")
            os.makedirs(createdPath, exist_ok=True)
            full_slide_name = str(image_index) + "_" + folder + "_" + os.path.splitext(image['name'])[0]
            print(f"\n\nNow saving '{full_slide_name}' into a spreasheet...")
            wb = Workbook()
            ws = wb.active

            # Raw filtered date
            ws.merge_cells('A1:H1')
            ws['A1'] = full_slide_name
            ws['A1'].font = Font(bold=True)
            ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
            ws['A1'].fill = PatternFill(start_color='009797', end_color='009797', patternType='solid')
            rows = dataframe_to_rows(raw_data, index=False, header=True)
            start_row = 2
            start_column = 1
            for r_idx, row in enumerate(rows, start_row):
                for c_idx, value in enumerate(row, start_column):
                    cell = ws.cell(row=r_idx, column=c_idx, value=value)

                    if r_idx == start_row:
                        column_letter = get_column_letter(c_idx)
                        column_width = len(str(value)) + 1
                        ws.column_dimensions[column_letter].width = column_width
                    elif r_idx % 2 != 0:
                        cell.fill = PatternFill(start_color='99d5d5', end_color='99d5d5', fill_type='solid')

            # Outlier data P3
            ws.merge_cells('J1:Q1')
            ws['J1'] = "Outliers of Phase 3"
            ws['J1'].font = Font(bold=True)
            ws['J1'].alignment = Alignment(horizontal='center', vertical='center')
            ws['J1'].fill = PatternFill(start_color='7d8a89', end_color='7d8a89', patternType='solid')
            if not raw_data_outlier_P3.empty:
                rows = dataframe_to_rows(raw_data_outlier_P3, index=False, header=True)
                start_row = 2
                start_column = 10
                for r_idx, row in enumerate(rows, start_row):
                    for c_idx, value in enumerate(row, start_column):
                        cell = ws.cell(row=r_idx, column=c_idx, value=value)

                        if r_idx == start_row:
                            column_letter = get_column_letter(c_idx)
                            column_width = len(str(value)) + 1
                            ws.column_dimensions[column_letter].width = column_width
                        elif r_idx % 2 != 0:
                            cell.fill = PatternFill(start_color='c4c3bf', end_color='c4c3bf', fill_type='solid')


            # Outlier data P2
            ws.merge_cells('S1:Z1')
            ws['S1'] = "Outliers of Phase 2"
            ws['S1'].font = Font(bold=True)
            ws['S1'].alignment = Alignment(horizontal='center', vertical='center')
            ws['S1'].fill = PatternFill(start_color='7d8a89', end_color='7d8a89', patternType='solid')
            if not raw_data_outlier_P2.empty:
                rows = dataframe_to_rows(raw_data_outlier_P2, index=False, header=True)
                start_row = 2
                start_column = 19
                for r_idx, row in enumerate(rows, start_row):
                    for c_idx, value in enumerate(row, start_column):
                        cell = ws.cell(row=r_idx, column=c_idx, value=value)

                        if r_idx == start_row:
                            column_letter = get_column_letter(c_idx)
                            column_width = len(str(value)) + 1
                            ws.column_dimensions[column_letter].width = column_width
                        elif r_idx % 2 != 0:
                            cell.fill = PatternFill(start_color='c4c3bf', end_color='c4c3bf', fill_type='solid')


            # Normal Distribution Test
            ws.merge_cells('AB1:AH1')
            ws['AB1'] = "Normal Distribution Test"
            ws['AB1'].font = Font(bold=True)
            ws['AB1'].alignment = Alignment(horizontal='center', vertical='center')
            ws['AB1'].fill = PatternFill(start_color='97ce7d', end_color='97ce7d', patternType='solid')
            # Stages
            ws['AB3'] = "Before P3:"
            ws['AB3'].font = Font(bold=True)
            ws.column_dimensions['AB'].width = len("Before P3:") + 1
            ws['AB3'].fill = PatternFill(start_color='d3ecbc', end_color='d3ecbc', patternType='solid')
            ws['AB4'] = "After P3:"
            ws['AB4'].font = Font(bold=True)
            # Mean
            ws['AC2'] = "Mean:"
            ws['AC3'] = test_before_results['Mean']
            ws['AC3'].fill = PatternFill(start_color='d3ecbc', end_color='d3ecbc', patternType='solid')
            ws['AC4'] = test_after_results['Mean']
            # Median
            ws['AD2'] = "Median:"
            ws['AD3'] = test_before_results['Median']
            ws['AD3'].fill = PatternFill(start_color='d3ecbc', end_color='d3ecbc', patternType='solid')
            ws['AD4'] = test_after_results['Median']
            # Mode
            ws['AE2'] = "Mode:"
            ws['AE3'] = test_before_results['Mode']
            ws['AE3'].fill = PatternFill(start_color='d3ecbc', end_color='d3ecbc', patternType='solid')
            ws['AE4'] = test_after_results['Mode']
            # Skewness Statistic
            ws['AF2'] = "Skewness Statistic:"
            ws.column_dimensions['AF'].width = len("Skewness Statistic:") + 1
            ws['AF3'] = test_before_results['Skewness Statistic']
            ws['AF3'].fill = PatternFill(start_color='d3ecbc', end_color='d3ecbc', patternType='solid')
            ws['AF4'] = test_after_results['Skewness Statistic']
            # Shapiro-Wilk Test Statistic
            ws['AG2'] = "Shapiro-Wilk Test Statistic:"
            ws.column_dimensions['AG'].width = len("Shapiro-Wilk Test Statistic:") + 1
            ws['AG3'] = test_before_results['Sharpio Statistic']
            ws['AG3'].fill = PatternFill(start_color='d3ecbc', end_color='d3ecbc', patternType='solid')
            ws['AG4'] = test_after_results['Sharpio Statistic']
            # P-value (Shapiro)
            ws['AH2'] = "P-value (Shapiro):"
            ws.column_dimensions['AH'].width = len("P-value (Shapiro):") + 1
            ws['AH3'] = test_before_results['Sharpio P-Value']
            ws['AH3'].fill = PatternFill(start_color='d3ecbc', end_color='d3ecbc', patternType='solid')
            ws['AH4'] = test_after_results['Sharpio P-Value']


            # Total Cell Count
            ws.merge_cells('AB6:AE6')
            ws['AB6'] = "Total Cell Count After Each Phase"
            ws['AB6'].font = Font(bold=True)
            ws['AB6'].alignment = Alignment(horizontal='center', vertical='center')
            ws['AB6'].fill = PatternFill(start_color='f9d52b', end_color='f9d52b', patternType='solid')
            # Orginal Total
            ws['AB7'] = "Orginal:"
            ws['AB8'] = total_OG_cells
            ws['AB8'].fill = PatternFill(start_color='ffe599', end_color='ffe599', patternType='solid')
            # P1 Total
            ws['AC7'] = "Phase 1 (P1):"
            ws.column_dimensions['AC'].width = len("Phase 1 (P1):") + 1
            ws['AC8'] = total_filtered_cells_P1
            ws['AC8'].fill = PatternFill(start_color='ffe599', end_color='ffe599', patternType='solid')
            # P2 Total
            ws['AD7'] = "Phase 2 (P2):"
            ws.column_dimensions['AD'].width = len("Phase 2 (P2):") + 1
            ws['AD8'] = total_filtered_cells_P2
            ws['AD8'].fill = PatternFill(start_color='ffe599', end_color='ffe599', patternType='solid')
            # P3 Total
            ws['AE7'] = "Phase 3 (P3):"
            ws.column_dimensions['AE'].width = len("Phase 3 (P3):") + 1
            ws['AE8'] = total_filtered_cells_P3
            ws['AE8'].fill = PatternFill(start_color='ffe599', end_color='ffe599', patternType='solid')
            
            if images_in_sheets:
                # Normal Distribution Figure
                img1 = XLImage(normal_distribution_bufferFig)
                img1.width = 1920
                img1.height = 1080
                ws.add_image(img1,'AJ1')
                
                # Image Processing Figure
                img2 = XLImage(img_processing_bufferFig)
                img2.width = 1920
                img2.height = 1080
                ws.add_image(img2,'AJ60')
            

            full_slide_name = full_slide_name + ".xlsx"
            file_path = os.path.join(createdPath,full_slide_name)
            wb.save(file_path)
            now = datetime.now()
            number_of_slides_done = number_of_slides_done + 1
            percentage_done_for_sample = round(number_of_slides_done/number_of_slides_to_analyze * 100, 2)
            print(f"Save has been completed at {now.strftime('%H:%M:%S')}")
            print(f"Sample Completion at: {percentage_done_for_sample}%...\n")
        
    # At the end of processing horizontal type images, dir changes back to analysis folder
    os.chdir(analysis_folder)